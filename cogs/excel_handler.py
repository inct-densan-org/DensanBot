import discord
from discord import app_commands
from discord.ext import commands
import os
import io
import zipfile
import datetime
import openpyxl

# 共通関数と定数をutilsからインポート
from .utils import get_wareki, get_excel_filename_for_month

class ExcelHandlerCog(commands.Cog):
    def __init__(self, bot: commands.Bot):
        self.bot = bot

    excel = app_commands.Group(name="excel", description="Excelファイルに関する操作")

    @excel.command(name="export", description="指定された年月のExcelファイルを出力します。")
    @app_commands.describe(year="西暦年 (省略時は今年)", month="月 (省略時は今月)", as_zip="[デフォルト: True] zip形式で出力")
    async def export_excel(self, interaction: discord.Interaction, year: int = None, month: int = None, as_zip: bool = True):
        try:
            await interaction.response.defer(ephemeral=False, thinking=True)
            today = datetime.date.today()
            target_year = year if year is not None else today.year
            target_month = month if month is not None else today.month
            target_filename = get_excel_filename_for_month(target_year, target_month)
            if not os.path.exists(target_filename):
                await interaction.followup.send(f"エラー: {target_year}年{target_month}月のExcelファイルは見つかりません。", ephemeral=True)
                return
            
            workbook = openpyxl.load_workbook(target_filename)
            if "活動計画書" in workbook.sheetnames: workbook["活動計画書"]["H1"] = f"{today.month}月{today.day}日"
            if "活動報告書" in workbook.sheetnames:
                wareki_year_str, _ = get_wareki(today.year)
                workbook["活動報告書"]["A1"] = f"{wareki_year_str}{today.month}月{today.day}日"
            workbook.save(target_filename)

            if as_zip:
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(target_filename, arcname=os.path.basename(target_filename))
                zip_buffer.seek(0)
                output_filename = f"report_{target_year}{target_month:02}.zip"
                file_to_send = discord.File(zip_buffer, filename=output_filename)
                await interaction.followup.send(f"{target_year}年{target_month}月のExcelファイルをzip形式で出力しました。", file=file_to_send)
            else:
                file_to_send = discord.File(target_filename, filename=os.path.basename(target_filename))
                await interaction.followup.send(f"{target_year}年{target_month}月のExcelファイルを出力しました。(ファイル名が文字化けする可能性があります)", file=file_to_send)
        except Exception as e:
            print(f"An error occurred in export_excel: {e}")
            await interaction.followup.send(f"申し訳ありません、エラーが発生しました: {e}", ephemeral=True)

async def setup(bot: commands.Bot):
    await bot.add_cog(ExcelHandlerCog(bot))
